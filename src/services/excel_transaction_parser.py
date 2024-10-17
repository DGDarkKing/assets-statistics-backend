import traceback
from typing import IO, io, Self

from openpyxl.cell import Cell
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, model_validator

from schemas.transaction import Transaction
from services.interfaces.i_file_transaction_parser import IFileTransactionParser


class ParseData(BaseModel):
    row: int
    symbol_colm: int
    direction_colm: int
    timestamp_colm: int

    amount_usdt_colm: int | None
    price_colm: int | None
    amount_coin_colm: int | None

    @model_validator(mode='after')
    def _validator(self) -> Self:
        if any(filter(lambda x: (None, None) == x,
                      (
                              (self.amount_usdt_colm, self.amount_coin_colm),
                              (self.amount_usdt_colm, self.price_colm),
                              (self.amount_coin_colm, self.price_colm),
                      ))):
            raise ValueError('Not found at least one pair: '
                             '"Amount usdt"-"Amount coin"; '
                             '"Amount usdt"-"Price"; '
                             '"Amount coin"-"Price".')
        return self


class ExcelTransactionParser(IFileTransactionParser):
    REQUIRED_FIELDS = [
        'symbol',
        'direction',
        'timestamp'
    ]
    AT_LEAST_ONE_REQUIRED_FIELDS_BY_RULES = {
        'price': {'amount usdt', 'amount coin'},
        'amount usdt': {'price', 'amount coin'},
        'amount coin': {'amount usdt', 'price'},
    }

    def parse(self, file: str | IO) -> list[Transaction]:
        with open(file, 'rb') as fbr:
            file = fbr.read()
            file = io.BytesIO(file)

        workbook: Workbook = load_workbook(file, read_only=True)
        sheet: Worksheet = workbook.active
        parse_data = self._find_columns(sheet)

        transactions: list[Transaction] = []
        while True:
            parse_data.row += 1
            if sheet.cell(row=parse_data.row, column=parse_data.symbol_colm).value is None:
                break

            amount_usdt: float = None
            amount_coin: float = None
            price: float = None
            try:
                if parse_data.amount_usdt_colm:
                    amount_usdt = float(sheet.cell(row=parse_data.row, column=parse_data.amount_usdt_colm).value)
            except:
                print(traceback.format_exc())
            try:
                if parse_data.amount_coin_colm:
                    amount_coin = float(sheet.cell(row=parse_data.row, column=parse_data.amount_coin_colm).value)
            except:
                print(traceback.format_exc())

            if amount_coin is None and amount_usdt is None:
                raise ValueError(f'Row {parse_data.row}: '
                                 f'failed parse "Amount coin"({parse_data.amount_coin_colm}) '
                                 f'and "Amount usdt"({parse_data.amount_usdt_colm})')

            if amount_usdt is None or amount_coin is None:
                try:
                    price = float(sheet.cell(row=parse_data.row, column=parse_data.price_colm).value)
                except:
                    print(traceback.format_exc())
                if not price:
                    raise ValueError(f'Row {parse_data.row}: '
                                     f'failed parse "Price"({parse_data.price_colm})')

                if amount_usdt is None:
                    amount_usdt = amount_coin * price
                elif amount_coin is None:
                    amount_coin = amount_usdt / price

            transaction = Transaction(
                symbol=sheet.cell(row=parse_data.row, column=parse_data.symbol_colm).value,
                direction=sheet.cell(row=parse_data.row, column=parse_data.direction_colm).value,
                amount_usdt=amount_usdt,
                amount_coin=amount_coin,
                dt=sheet.cell(parse_data.row, column=parse_data.timestamp_colm).value
            )
            transactions.append(transaction)

        return transactions

    def _find_columns(
            self,
            work_sheet: Worksheet
    ) -> ParseData:
        num_params = 6
        symbol_colm, direction_colm, timestamp_colm, amount_usdt_colm, price_colm, amount_coin_colm = (
                (None,) * num_params)

        row, column = 1, 1
        max_row = 10
        cell: Cell = work_sheet.cell(row=row, column=column)
        while not cell.value and row <= max_row:
            row += 1
            cell: Cell = work_sheet.cell(row=row, column=column)

        if not cell.value:
            raise ValueError('Wrong document style')

        found_params_counter = 0
        while cell.value and found_params_counter < num_params:
            match cell.value.lower():
                case 'symbol':
                    if not symbol_colm:
                        symbol_colm = column
                case 'direction':
                    if not direction_colm:
                        direction_colm = column
                case 'timestamp':
                    if not timestamp_colm:
                        timestamp_colm = column

                case 'price':
                    if not price_colm:
                        price_colm = column
                case x if x in ('amount usdt', 'amount_usdt'):
                    if not amount_usdt_colm:
                        amount_usdt_colm = column
                case x if x in ('amount coin', 'amount_coin'):
                    if not amount_coin_colm:
                        amount_coin_colm = column

            column += 1
            cell: Cell = work_sheet.cell(row=row, column=column)

        return ParseData(
            row=row,
            symbol_colm=symbol_colm,
            direction_colm=direction_colm,
            timestamp_colm=timestamp_colm,
            amount_usdt_colm=amount_usdt_colm,
            price_colm=price_colm,
            amount_coin_colm=amount_coin_colm
        )
